"""
Inserts Section 3 charts from the charts/ folder into the 'Chart' sheet
of HW Q3 Problem_Agency MBS Research 2.xlsx.
Charts are laid out in a 2-column grid with labels.
"""

import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = r"c:\Users\nduta\OneDrive\Desktop\Projects\mbs"
Q3_FILE = os.path.join(BASE_DIR, "HW Q3 Problem_Agency MBS Research 2.xlsx")
CHARTS_DIR = os.path.join(BASE_DIR, "charts")

# Charts relevant to Section 3 (Q3), in display order
SECTION3_CHARTS = [
    ("q3_1_issuance.png",       "Q3.1 – Issuance by Year & Loan Purpose"),
    ("q3_3a_balance.png",       "Q3.3a – Current Balance Over Time"),
    ("q3_3b_cpr.png",           "Q3.3b – Historical CPR Over Time"),
    ("q3_4_wac_cpr.png",        "Q3.4 – WAC vs CPR (Line Chart)"),
    ("q3_4_wac_cpr_scatter.png","Q3.4 – WAC vs CPR (Scatter Chart)"),
]

# Layout constants (in Excel row/column units)
# Each image block: title row + image rows
# We'll use a 2-column layout: columns A-H (left) and I-P (right)
# Row spacing: 1 title row + 20 image rows per image

COLS_LEFT  = "A"   # anchor column for left images
COLS_RIGHT = "J"   # anchor column for right images
TITLE_HEIGHT = 25  # points
IMG_ROWS = 20      # rows per image
IMG_WIDTH_PX = 480
IMG_HEIGHT_PX = 300

def set_title_cell(ws, cell_ref, title):
    cell = ws[cell_ref]
    cell.value = title
    cell.font = Font(bold=True, size=12, color="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill("solid", fgColor="D6E4F0")

def row_to_start(img_index):
    """Return the starting row for image block (0-indexed)."""
    block = img_index // 2      # which row-pair
    start_row = 1 + block * (IMG_ROWS + 2)   # +2: title + gap
    return start_row

def main():
    wb = load_workbook(Q3_FILE)

    # Create or clear the Charts sheet
    if "Section 3 Charts" in wb.sheetnames:
        del wb["Section 3 Charts"]

    ws = wb.create_sheet("Section 3 Charts")

    # Set column widths for a neat layout
    for col in range(1, 20):
        ws.column_dimensions[get_column_letter(col)].width = 8

    # Set row heights
    for row in range(1, 200):
        ws.row_dimensions[row].height = 15

    for idx, (fname, label) in enumerate(SECTION3_CHARTS):
        img_path = os.path.join(CHARTS_DIR, fname)
        if not os.path.exists(img_path):
            print(f"  [SKIP] {fname} not found.")
            continue

        block_row = idx // 2     # which vertical block (0, 1, 2 …)
        col_left  = idx % 2 == 0 # even idx → left column, odd → right

        # Calculate row for title and image anchor
        title_row = 1 + block_row * (IMG_ROWS + 3)   # 3 = title + 2 gap rows
        img_row   = title_row + 1

        if col_left:
            title_col = "A"
            img_anchor_col = "A"
            title_cell = f"A{title_row}"
        else:
            title_col = "J"
            img_anchor_col = "J"
            title_cell = f"J{title_row}"

        # Merge title cells across span
        if col_left:
            merge_end = f"I{title_row}"
        else:
            merge_end = f"R{title_row}"

        try:
            ws.merge_cells(f"{title_cell}:{merge_end}")
        except Exception:
            pass

        # Set the title
        set_title_cell(ws, title_cell, label)

        # Add the image
        img = XLImage(img_path)
        img.width  = IMG_WIDTH_PX
        img.height = IMG_HEIGHT_PX
        img.anchor = f"{img_anchor_col}{img_row}"
        ws.add_image(img)

        print(f"  [OK] Added '{label}' at {img_anchor_col}{img_row}")

    # Also make the existing Chart sheet the active one
    wb.active = wb["Section 3 Charts"]

    wb.save(Q3_FILE)
    print(f"\nSaved to: {Q3_FILE}")

if __name__ == "__main__":
    main()
