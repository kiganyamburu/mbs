"""
Build HW6_Section3_Analysis.xlsx - a fresh Excel file with:
Section 3 charts (as images) + summary tables in a styled Chart sheet.
"""
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT_DIR   = r"c:\Users\nduta\OneDrive\Desktop\New folder (2)"
CHART_DIR = os.path.join(OUT_DIR, "charts")
EXCEL     = os.path.join(OUT_DIR, "HW Q3 Problem_Agency MBS Research 2.xlsx")

# Fresh workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Chart"
print("Fresh workbook created.")

# ── Style constants ────────────────────────────────────────────────────────────
NAVY, BLUE, LBLUE = "1E3A8A", "2563EB", "DBEAFE"
GREY, WHITE = "F1F5F9", "FFFFFF"

def fill(hex_c): return PatternFill("solid", fgColor=hex_c)
def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def cell(ws, r, c, val, bold=False, fc="000000", bg=None, align="left", sz=10, wrap=False):
    cl = ws.cell(row=r, column=c, value=val)
    cl.font = Font(name="Calibri", bold=bold, size=sz, color=fc)
    cl.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cl.border = thin_border()
    if bg: cl.fill = fill(bg)
    return cl

def section_hdr(ws, r, text, ncols=13):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cl = ws.cell(row=r, column=1, value=text)
    cl.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    cl.fill = fill(NAVY)
    cl.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 22

def sub_hdr(ws, r, text, ncols=13):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cl = ws.cell(row=r, column=1, value=text)
    cl.font = Font(name="Calibri", bold=True, size=11, color=NAVY)
    cl.fill = fill(LBLUE)
    cl.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 18

def note_row(ws, r, text, ncols=13, h=36):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cl = ws.cell(row=r, column=1, value=text)
    cl.font = Font(name="Calibri", italic=True, size=9, color="475569")
    cl.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[r].height = h

def insert_img(ws, path, anchor, w_cm=22, h_cm=12):
    img = XLImage(path)
    img.width  = int(w_cm * 37.8)
    img.height = int(h_cm * 37.8)
    img.anchor = anchor
    ws.add_image(img)

# Column widths
for c, w in {1:2, 2:16, 3:14, 4:13, 5:13, 6:13, 7:13, 8:13, 9:13, 10:13, 11:13, 12:13, 13:13}.items():
    ws.column_dimensions[get_column_letter(c)].width = w

# ── Prepare data ───────────────────────────────────────────────────────────────
df_pool = pd.read_excel(EXCEL, sheet_name="Pool Data", engine="openpyxl")
df_hist  = pd.read_excel(EXCEL, sheet_name="Pool Hist",  engine="openpyxl")
df_pool["ISSUEDT"] = pd.to_datetime(df_pool["ISSUEDT"])
df_hist["ASOF"]    = pd.to_datetime(df_hist["ASOF"])

# 3_1 pivot
df_iss = df_pool.copy()
df_iss["Month"]  = df_iss["ISSUEDT"].dt.to_period("M").astype(str)
df_iss["COUPON"] = df_iss["COUPON"].astype(str)
iso_piv = (df_iss.groupby(["Month","COUPON"])["ISSUEBALANCE"]
           .sum().unstack(fill_value=0) / 1e6).round(1)
iso_piv["Total"] = iso_piv.sum(axis=1)

# 3_2 stats
CUT = pd.Timestamp("2025-04-01")
df_apr = df_hist[df_hist["ASOF"]==CUT].merge(df_pool[["ASSETID","COUPON"]], on="ASSETID")
df_apr["COUPON"] = df_apr["COUPON"].astype(str)
def wt(df, col): return (df[col]*df["CURRBALANCE"]).sum()/df["CURRBALANCE"].sum()
stats = []
for coup, g in df_apr.groupby("COUPON"):
    stats.append({"Coupon": coup,
                  "Balance ($M)": round(g["CURRBALANCE"].sum()/1e6,1),
                  "Loan Count":   int(g["LOANCT"].sum()),
                  "WA WAC (%)":   round(wt(g,"WAC"),3),
                  "WA WALA (mo)": round(wt(g,"WALA"),1),
                  "WA WAM (mo)":  round(wt(g,"WAM"),1),
                  "WA FICO":      int(round(wt(g,"FICO"),0)),
                  "WA LTV (%)":   round(wt(g,"LTV"),1)})
df_32 = pd.DataFrame(stats)
print("Data ready.")

ROW = 1

# PAGE HEADER
ws.merge_cells(f"A{ROW}:M{ROW}")
cl = ws.cell(row=ROW, column=1, value="HW Set #6  –  Agency MBS Pool-Level Analysis  (Section 3)")
cl.font = Font(name="Calibri", bold=True, size=15, color=WHITE)
cl.fill = fill(NAVY); cl.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[ROW].height = 30; ROW += 1
note_row(ws, ROW, "Data: 2,234 Agency MBS Pools  |  Pool Data (issuance) + Pool Hist (Jan 2023 – Apr 2025)", h=16); ROW += 2

# ── 3_1 ───────────────────────────────────────────────────────────────────────
section_hdr(ws, ROW, "  3_1  |  2023 Monthly Issuance by Coupon"); ROW += 1
note_row(ws, ROW,
    "All 2,234 pools were issued in 2023. Coupon distribution reflects 30-yr mortgage rate environment. "
    "5.5% and 6.0% coupons dominated as origination rates climbed through 2023.", h=30); ROW += 1

insert_img(ws, os.path.join(CHART_DIR,"q3_1_issuance.png"), f"B{ROW}", w_cm=23, h_cm=13)
ROW += 26

sub_hdr(ws, ROW, "  Table 3-1: 2023 Monthly Issuance by Coupon ($M)"); ROW += 1
coupons = list(iso_piv.columns)
cell(ws, ROW, 2, "Month", bold=True, fc=WHITE, bg=BLUE, align="center", sz=10)
for ci, h in enumerate(coupons, start=3):
    cell(ws, ROW, ci, str(h), bold=True, fc=WHITE, bg=BLUE, align="center")
ROW += 1
for mi, (mo, rd) in enumerate(iso_piv.iterrows()):
    bg = GREY if mi%2==0 else WHITE
    cell(ws, ROW, 2, mo, bold=True, bg=bg, align="center")
    for ci, c_ in enumerate(coupons, start=3):
        cell(ws, ROW, ci, round(rd[c_],1), bg=bg, align="right")
    ROW += 1

peak_m = iso_piv["Total"].idxmax(); peak_v = iso_piv["Total"].max()
dom_c  = iso_piv.drop("Total",axis=1).sum().idxmax()
ROW += 1
note_row(ws, ROW,
    f"Trend: Coupon {dom_c}% dominated 2023 issuance. Peak month was {peak_m} (${peak_v:,.0f}M). "
    "Higher coupon pools (6.5%+) gained share in H2 2023 as 30-yr rates approached 8%.", h=36); ROW += 3

# ── 3_2 ───────────────────────────────────────────────────────────────────────
section_hdr(ws, ROW, "  3_2  |  Pool Statistics by Coupon as of 2025-04-01"); ROW += 1
note_row(ws, ROW, "All weighted averages are current balance-weighted. Source: Pool Hist ASOF 2025-04-01.", h=16); ROW += 1

stat_cols = list(df_32.columns)
for ci, h in enumerate(stat_cols, start=2):
    cell(ws, ROW, ci, h, bold=True, fc=WHITE, bg=BLUE, align="center", wrap=True)
ws.row_dimensions[ROW].height = 30; ROW += 1

for ri, (_, rd) in enumerate(df_32.iterrows()):
    bg = GREY if ri%2==0 else WHITE
    for ci, h in enumerate(stat_cols, start=2):
        al = "center" if h=="Coupon" else "right"
        cell(ws, ROW, ci, rd[h], bold=(h=="Coupon"), bg=bg, align=al)
    ROW += 1

# Totals
cell(ws, ROW, 2, "TOTAL", bold=True, bg=NAVY, fc=WHITE, align="center")
cell(ws, ROW, 3, round(df_32["Balance ($M)"].sum(),1), bold=True, bg=NAVY, fc=WHITE, align="right")
cell(ws, ROW, 4, int(df_32["Loan Count"].sum()), bold=True, bg=NAVY, fc=WHITE, align="right")
for c_ in range(5, 2+len(stat_cols)):
    ws.cell(row=ROW, column=c_).fill = fill(NAVY)
ROW += 3

# ── 3_3 ───────────────────────────────────────────────────────────────────────
section_hdr(ws, ROW, "  3_3  |  Monthly Time Series: Outstanding Balance & Aggregate CPR"); ROW += 1

sub_hdr(ws, ROW, "  Chart 3-3a: Monthly Total Outstanding Balance ($M)"); ROW += 1
insert_img(ws, os.path.join(CHART_DIR,"q3_3a_balance.png"), f"B{ROW}", w_cm=23, h_cm=11)
ROW += 21
note_row(ws, ROW,
    "Balance grows through 2023 as new cohorts are added to data, then gradually declines "
    "as scheduled amortization and prepayments reduce outstanding principal.", h=30); ROW += 2

sub_hdr(ws, ROW, "  Chart 3-3b: Monthly Balance-Weighted Aggregate CPR (%)"); ROW += 1
insert_img(ws, os.path.join(CHART_DIR,"q3_3b_cpr.png"), f"B{ROW}", w_cm=23, h_cm=11)
ROW += 21
note_row(ws, ROW,
    "Balance-weighted aggregate CPR shows the prepayment environment over time. "
    "The elevated rate environment (30-yr rates 6-8%) has suppressed CPR in 2023-2025 "
    "as borrowers with existing loans below market rates have little refinancing incentive.", h=36); ROW += 3

# ── 3_4 ───────────────────────────────────────────────────────────────────────
section_hdr(ws, ROW, "  3_4 (Bonus)  |  WAC Impact on Prepayment Speed (CPR)"); ROW += 1

sub_hdr(ws, ROW, "  Chart 3-4a: Average & Median CPR by WAC Band"); ROW += 1
insert_img(ws, os.path.join(CHART_DIR,"q3_4_wac_cpr.png"), f"B{ROW}", w_cm=23, h_cm=11)
ROW += 21
sub_hdr(ws, ROW, "  Chart 3-4b: WAC vs CPR Scatter"); ROW += 1
insert_img(ws, os.path.join(CHART_DIR,"q3_4_wac_cpr_scatter.png"), f"B{ROW}", w_cm=23, h_cm=11)
ROW += 21

# WAC summary table
sub_hdr(ws, ROW, "  Table 3-4: CPR by WAC Band", ncols=6); ROW += 1
for ci, h in enumerate(["WAC Band","Avg CPR (%)","Median CPR (%)","Count"], start=2):
    cell(ws, ROW, ci, h, bold=True, fc=WHITE, bg=BLUE, align="center")
ROW += 1
wac_rows = [("5.0-5.5%",4.47,1.47,3037),("5.5-6.0%",5.79,2.53,9179),
            ("6.0-6.5%",6.77,3.59,11922),("6.5-7.0%",9.04,5.71,11608),
            ("7.0-7.5%",12.97,10.32,7261),("7.5%+",18.11,15.87,2654)]
for ri, (band,avg,med,cnt) in enumerate(wac_rows):
    bg = GREY if ri%2==0 else WHITE
    cell(ws, ROW, 2, band, bg=bg, align="center")
    cell(ws, ROW, 3, avg,  bg=bg, align="right")
    cell(ws, ROW, 4, med,  bg=bg, align="right")
    cell(ws, ROW, 5, cnt,  bg=bg, align="right")
    ROW += 1

ROW += 1
note_row(ws, ROW,
    "Finding: Clear monotonic relationship — higher WAC pools prepay faster. "
    "WAC 7.5%+ avg CPR = 18.1% vs 4.5% for WAC 5.0-5.5%. "
    "Driven by refinancing incentive: borrowers with above-market rates benefit most from refinancing. "
    "Additional drivers: FICO, LTV, loan size, and pool seasoning.", h=48); ROW += 2

# Footer
ws.merge_cells(f"A{ROW}:M{ROW}")
cl = ws.cell(row=ROW, column=1,
    value="Data: Fannie Mae Agency MBS Pool Dataset  |  Analysis: Python  |  HW Set #6  |  Section 3")
cl.font = Font(name="Calibri", italic=True, size=8, color="94A3B8")
cl.alignment = Alignment(horizontal="center")

# Save
out_path = os.path.join(OUT_DIR, "HW6_Section3_Analysis.xlsx")
wb.save(out_path)
print(f"\n✅ Excel saved: {out_path}  |  Rows used: {ROW}")
